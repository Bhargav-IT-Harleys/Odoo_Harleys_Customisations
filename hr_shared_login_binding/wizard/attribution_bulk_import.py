"""UI front-end for the same resolve-then-apply pattern used by the
Odoo-shell bulk-linking scripts: upload a two-column file (functional
account email, employee email), review what resolves, then apply.
"""
import base64
import csv
import io
import re

from odoo import api, fields, models, _
from odoo.exceptions import UserError

TEMPLATE_HEADERS = ("Functional Account Email", "Employee Email")


def _split_emails(cell_value):
    """One cell may hold several comma/semicolon-separated emails (the
    natural shape of a "functional account -> its employees" list), each
    possibly followed by a parenthetical location tag like "(BKC)". Splits
    into individual, tag-stripped addresses.
    """
    if not cell_value:
        return []
    cell_value = re.sub(r'\([^)]*\)', '', cell_value)
    return [p.strip() for p in re.split(r'[\s,;]+', cell_value) if p.strip()]


class AttributionBulkImportLine(models.TransientModel):
    _name = 'hr.attribution.bulk.import.line'
    _description = "Employee Attribution Bulk Import - Row"

    wizard_id = fields.Many2one('hr.attribution.bulk.import', required=True, ondelete='cascade')
    functional_email = fields.Char(string="Functional Email")
    employee_email = fields.Char(string="Employee Email")
    # Editable, not just parse output: fixing an unresolved row (or adding
    # a new one by hand) is done by picking these directly, rather than
    # forcing a re-upload for one bad row.
    functional_user_id = fields.Many2one('res.users', string="Functional Account")
    employee_id = fields.Many2one('hr.employee', string="Employee")
    status = fields.Selection(
        [('resolved', "Resolved"), ('unresolved', "Unresolved")],
        compute='_compute_status', store=True,
    )
    reason = fields.Char(readonly=True)

    @api.depends('functional_user_id', 'employee_id')
    def _compute_status(self):
        for line in self:
            line.status = 'resolved' if (line.functional_user_id and line.employee_id) else 'unresolved'
            if line.status == 'resolved':
                line.reason = False

    @api.onchange('functional_user_id')
    def _onchange_functional_user_id(self):
        if self.functional_user_id:
            self.functional_email = self.functional_user_id.login

    @api.onchange('employee_id')
    def _onchange_employee_id(self):
        if self.employee_id:
            self.employee_email = (
                self.employee_id.work_email
                or self.employee_id.private_email
                or self.employee_id.user_id.login
            )


class AttributionBulkImport(models.TransientModel):
    _name = 'hr.attribution.bulk.import'
    _description = "Employee Attribution Bulk Import"

    state = fields.Selection([
        ('draft', "Upload"),
        ('validated', "Validated"),
        ('done', "Done"),
    ], default='draft', required=True)
    import_file = fields.Binary(string="Import File (.xlsx or .csv)")
    import_filename = fields.Char()
    template_file = fields.Binary(readonly=True)
    template_filename = fields.Char(default="employee_attribution_template.xlsx")
    result_file = fields.Binary(readonly=True)
    result_filename = fields.Char(default="employee_attribution_result.xlsx")
    line_ids = fields.One2many('hr.attribution.bulk.import.line', 'wizard_id')
    resolved_count = fields.Integer(compute='_compute_counts')
    unresolved_count = fields.Integer(compute='_compute_counts')

    @api.depends('line_ids.status')
    def _compute_counts(self):
        for wizard in self:
            wizard.resolved_count = len(wizard.line_ids.filtered(lambda l: l.status == 'resolved'))
            wizard.unresolved_count = len(wizard.line_ids.filtered(lambda l: l.status == 'unresolved'))

    def action_generate_template(self):
        self.ensure_one()
        import xlsxwriter  # noqa: PLC0415 - optional dependency, only needed here
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        sheet = workbook.add_worksheet('Attribution')
        bold = workbook.add_format({'bold': True})
        for col, header in enumerate(TEMPLATE_HEADERS):
            sheet.write(0, col, header, bold)
        sheet.set_column(0, 1, 35)
        workbook.close()
        output.seek(0)
        self.template_file = base64.b64encode(output.read())
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _read_rows(self):
        if not self.import_file:
            raise UserError(_("Upload a file first."))
        content = base64.b64decode(self.import_file)
        filename = (self.import_filename or '').lower()
        raw_rows = []
        if filename.endswith('.csv'):
            text = content.decode('utf-8-sig')
            for i, row in enumerate(csv.reader(io.StringIO(text))):
                if i == 0 or not row or not row[0]:
                    continue
                raw_rows.append((row[0].strip(), row[1] if len(row) > 1 else ''))
        else:
            import openpyxl  # noqa: PLC0415 - optional dependency, only needed here
            book = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            for i, row in enumerate(book.active.iter_rows(values_only=True)):
                if i == 0 or not row or not row[0]:
                    continue
                raw_rows.append((str(row[0]).strip(), str(row[1]) if len(row) > 1 and row[1] else ''))
        # One functional account per row, but its employee cell may list
        # several addresses - expand into one (functional, employee) pair
        # per address rather than requiring the file pre-flattened.
        rows = []
        for functional_email, employee_cell in raw_rows:
            if not functional_email:
                continue
            for employee_email in _split_emails(employee_cell):
                rows.append((functional_email, employee_email))
        return rows

    def action_validate(self):
        self.ensure_one()
        rows = self._read_rows()
        self.line_ids.unlink()
        User = self.env['res.users']
        Employee = self.env['hr.employee']
        lines = []
        for functional_email, employee_email in rows:
            func_users = User.search([('login', '=ilike', functional_email)])
            emp_user = User.search([('login', '=ilike', employee_email)])
            emp = Employee.browse()
            if len(emp_user) == 1:
                emp = Employee.search([('user_id', '=', emp_user.id)])
            if not emp:
                emp = Employee.search([
                    '|', ('work_email', '=ilike', employee_email),
                    ('private_email', '=ilike', employee_email),
                ])
            # status isn't set here - it's computed from functional_user_id/
            # employee_id, so it stays consistent whether a row comes from
            # this parse or is added/fixed by hand afterwards.
            vals = {'functional_email': functional_email, 'employee_email': employee_email}
            if len(func_users) != 1:
                vals['reason'] = _("Functional account: %s match(es)") % len(func_users)
            elif len(emp) != 1:
                vals['reason'] = _("Employee: %s match(es)") % len(emp)
            else:
                vals.update(functional_user_id=func_users.id, employee_id=emp.id)
            lines.append((0, 0, vals))
        self.write({'line_ids': lines, 'state': 'validated'})
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_discard_unresolved(self):
        self.ensure_one()
        self.line_ids.filtered(lambda l: l.status == 'unresolved').unlink()
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_download_result(self):
        self.ensure_one()
        import xlsxwriter  # noqa: PLC0415 - optional dependency, only needed here
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        sheet = workbook.add_worksheet('Result')
        bold = workbook.add_format({'bold': True})
        headers = ("Functional Account Email", "Employee Email", "Functional Account",
                   "Employee", "Status", "Reason")
        for col, header in enumerate(headers):
            sheet.write(0, col, header, bold)
        for row, line in enumerate(self.line_ids, start=1):
            sheet.write(row, 0, line.functional_email or '')
            sheet.write(row, 1, line.employee_email or '')
            sheet.write(row, 2, line.functional_user_id.login or '')
            sheet.write(row, 3, line.employee_id.name or '')
            sheet.write(row, 4, line.status or '')
            sheet.write(row, 5, line.reason or '')
        sheet.set_column(0, 5, 30)
        workbook.close()
        output.seek(0)
        self.result_file = base64.b64encode(output.read())
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_apply(self):
        self.ensure_one()
        if self.state != 'validated':
            raise UserError(_("Validate the file before applying it."))
        resolved = self.line_ids.filtered(lambda l: l.status == 'resolved')
        if not resolved:
            raise UserError(_("Nothing resolved to apply."))
        by_user = {}
        for line in resolved:
            by_user.setdefault(line.functional_user_id, self.env['hr.employee'])
            by_user[line.functional_user_id] |= line.employee_id
        for func_user, employees in by_user.items():
            # Add rather than replace, so a repeated/updated import doesn't
            # wipe authorizations set by an earlier import or by hand. Also
            # flags the account as shared - authorized_employee_ids is
            # hidden in the Users view otherwise (invisible="not
            # is_shared_login"), so without this the import looks like it
            # did nothing even though the data is there.
            func_user.write({
                'is_shared_login': True,
                'authorized_employee_ids': [(4, eid) for eid in employees.ids],
            })
        self.state = 'done'
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Employee attribution updated"),
                'message': _(
                    "%(accounts)s functional account(s), %(employees)s employee(s) authorized.",
                    accounts=len(by_user), employees=len(resolved),
                ),
                'sticky': False,
            },
        }
